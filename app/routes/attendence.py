from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    AttendanceSession,
    AttendanceRecord,
    Student,
    Teacher
)

import csv
import io

from fastapi.responses import StreamingResponse

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    UploadFile,
)

from sqlalchemy.orm import Session

from app.database import get_db
from app.utility import get_current_teacher

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.models import (
    AttendanceRecord,
    AttendanceSession,
)

from app.services.attendance_engine import (
    AttendanceEngine,
)

from app.services.face_embedding_service import (
    FaceEmbeddingService,
)

from app.services.face_recognition_service import (
    FaceRecognitionService,
)


router = APIRouter(
    prefix="/api/attendance",
    tags=["Attendance"]
)


# ============================================================
# START ATTENDANCE SESSION
# ============================================================

@router.post(
    "/sessions",
    status_code=status.HTTP_201_CREATED,
)
def start_attendance_session(
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Check if THIS teacher already has an active session
    # --------------------------------------------------------

    active_session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.status == "ACTIVE",
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    if active_session:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "You already have an active "
                "attendance session"
            ),
        )

    # --------------------------------------------------------
    # 2. Get all active students
    # --------------------------------------------------------

    students = (
        db.query(Student)
        .filter(
            Student.is_active == True
        )
        .order_by(Student.id)
        .all()
    )

    if not students:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active students found",
        )

    # --------------------------------------------------------
    # 3. Create attendance session
    # --------------------------------------------------------

    session = AttendanceSession(
        teacher_id=current_teacher.id,
        status="ACTIVE",
        started_at=datetime.utcnow(),
    )

    db.add(session)

    # We need session.id before creating records
    db.flush()

    # --------------------------------------------------------
    # 4. Create ABSENT record for every student
    # --------------------------------------------------------

    attendance_records = []

    for student in students:

        record = AttendanceRecord(
            session_id=session.id,
            student_id=student.id,
            status="ABSENT",
        )

        attendance_records.append(record)

    db.add_all(attendance_records)

    # --------------------------------------------------------
    # 5. Save everything
    # --------------------------------------------------------

    db.commit()

    # Refresh session so we have latest database values
    db.refresh(session)

    # --------------------------------------------------------
    # 6. Return session information
    # --------------------------------------------------------

    return {
        "id": session.id,
        "status": session.status,
        "started_at": session.started_at,
        "total_students": len(students),
        "teacher": {
            "id": current_teacher.id,
            "name": current_teacher.name,
        },
    }

# ============================================================
# GET ACTIVE ATTENDANCE SESSION
# ============================================================

@router.get(
    "/sessions/active"
)
def get_active_attendance_session(
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find active session for CURRENT TEACHER
    # --------------------------------------------------------

    active_session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.status == "ACTIVE",
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. No active session for this teacher
    # --------------------------------------------------------

    if not active_session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No active attendance session found",
        )

    # --------------------------------------------------------
    # 3. Get attendance records
    # --------------------------------------------------------

    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == active_session.id
        )
        .all()
    )

    # --------------------------------------------------------
    # 4. Build student attendance list
    # --------------------------------------------------------

    students = []

    for record in records:

        students.append({
            "student_id": record.student.id,
            "name": record.student.name,
            "roll_number": record.student.roll_number,
            "email": record.student.email,
            "department": record.student.department,
            "semester": record.student.semester,
            "status": record.status,
            "marked_at": record.marked_at,
        })

    # --------------------------------------------------------
    # 5. Calculate statistics
    # --------------------------------------------------------

    total_students = len(students)

    present_students = sum(
        1
        for student in students
        if student["status"] == "PRESENT"
    )

    absent_students = sum(
        1
        for student in students
        if student["status"] == "ABSENT"
    )

    # --------------------------------------------------------
    # 6. Return active session
    # --------------------------------------------------------

    return {
        "session": {
            "id": active_session.id,
            "status": active_session.status,
            "started_at": active_session.started_at,
            "ended_at": active_session.ended_at,
        },

        "statistics": {
            "total_students": total_students,
            "present": present_students,
            "absent": absent_students,
        },

        "students": students,
    }


# ============================================================
# END / COMPLETE ATTENDANCE SESSION
# ============================================================

@router.patch(
    "/sessions/{session_id}/end"
)
def end_attendance_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find the session belonging to CURRENT TEACHER
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. Session does not exist / does not belong to teacher
    # --------------------------------------------------------

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # --------------------------------------------------------
    # 3. Check whether session is already completed
    # --------------------------------------------------------

    if session.status == "COMPLETED":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Attendance session is already completed",
        )

    # --------------------------------------------------------
    # 4. Only ACTIVE sessions can be completed
    # --------------------------------------------------------

    if session.status != "ACTIVE":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"Cannot complete session "
                f"with status '{session.status}'"
            ),
        )

    # --------------------------------------------------------
    # 5. Complete the session
    # --------------------------------------------------------

    session.status = "COMPLETED"
    session.ended_at = datetime.utcnow()

    db.commit()
    db.refresh(session)

    # --------------------------------------------------------
    # 6. Get attendance statistics
    # --------------------------------------------------------

    total_students = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id
        )
        .count()
    )

    present_students = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id,
            AttendanceRecord.status == "PRESENT",
        )
        .count()
    )

    absent_students = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id,
            AttendanceRecord.status == "ABSENT",
        )
        .count()
    )

    # --------------------------------------------------------
    # 7. Return completed session
    # --------------------------------------------------------

    return {
        "message": (
            "Attendance session "
            "completed successfully"
        ),

        "session": {
            "id": session.id,
            "status": session.status,
            "started_at": session.started_at,
            "ended_at": session.ended_at,
        },

        "teacher": {
            "id": current_teacher.id,
            "name": current_teacher.name,
        },

        "statistics": {
            "total_students": total_students,
            "present": present_students,
            "absent": absent_students,
        },
    }


# ============================================================
# MARK STUDENT PRESENT
# ============================================================

@router.post(
    "/sessions/{session_id}/mark-present"
)
def mark_student_present(
    session_id: int,
    student_id: int,
    db: Session = Depends(get_db)
):

    # --------------------------------------------------------
    # 1. Find the attendance session
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id
        )
        .first()
    )

    if not session:

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found"
        )

    # --------------------------------------------------------
    # 2. Make sure session is ACTIVE
    # --------------------------------------------------------

    if session.status != "ACTIVE":

        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Attendance session is not active"
        )

    # --------------------------------------------------------
    # 3. Find the student
    # --------------------------------------------------------

    student = (
        db.query(Student)
        .filter(
            Student.id == student_id,
            Student.is_active == True
        )
        .first()
    )

    if not student:

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Active student not found"
        )

    # --------------------------------------------------------
    # 4. Find student's attendance record
    # --------------------------------------------------------

    attendance_record = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id == session_id,
            AttendanceRecord.student_id == student_id
        )
        .first()
    )

    # --------------------------------------------------------
    # 5. Attendance record should already exist
    # --------------------------------------------------------

    if not attendance_record:

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance record not found for this student"
        )

    # --------------------------------------------------------
    # 6. Check if already present
    # --------------------------------------------------------

    if attendance_record.status == "PRESENT":

        return {
            "message": "Student is already marked present",

            "student": {
                "id": student.id,
                "name": student.name,
                "roll_number": student.roll_number
            },

            "attendance": {
                "status": attendance_record.status,
                "marked_at": attendance_record.marked_at
            }
        }

    # --------------------------------------------------------
    # 7. Mark student PRESENT
    # --------------------------------------------------------

    attendance_record.status = "PRESENT"
    attendance_record.marked_at = datetime.utcnow()

    db.commit()
    db.refresh(attendance_record)

    # --------------------------------------------------------
    # 8. Return result
    # --------------------------------------------------------

    return {
        "message": "Student marked present successfully",

        "student": {
            "id": student.id,
            "name": student.name,
            "roll_number": student.roll_number
        },

        "attendance": {
            "status": attendance_record.status,
            "marked_at": attendance_record.marked_at
        }
    }

@router.post("/sessions/{session_id}/recognize")
async def recognize_attendance(
    session_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):
    # ============================================================
    # 1. FIND ATTENDANCE SESSION
    # ============================================================
    session = (
    db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
      .first()
    )

    if session is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # ============================================================
    # 2. CHECK SESSION STATUS
    # ============================================================
    if session.status != "ACTIVE":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Attendance session is not active",
        )

    # ============================================================
    # 3. VALIDATE FILE
    # ============================================================
    allowed_types = {"image/jpeg", "image/png", "image/webp"}

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported image format",
        )

    # ============================================================
    # 4. READ IMAGE
    # ============================================================
    image_bytes = await file.read()

    if not image_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Empty image file",
        )

    # ============================================================
    # 5. LOAD KNOWN FACES FROM DATABASE
    # ============================================================
    known_embeddings, known_students = FaceRecognitionService.load_known_faces(db)

    if not known_embeddings:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No face embeddings available",
        )

    # ============================================================
    # 6. CREATE ATTENDANCE ENGINE
    # ============================================================
    engine = AttendanceEngine(
        known_embeddings=known_embeddings,
        known_students=known_students,
    )

    # ============================================================
    # 7. DETECT MULTIPLE FACES
    # ============================================================
    try:
        face_embeddings = FaceEmbeddingService.generate_embeddings(image_bytes)
    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Face processing failed: {str(error)}",
        )

    # ============================================================
    # 8. NO FACE DETECTED
    # ============================================================
    if not face_embeddings:
        return {
            "recognized": False,
            "faces_detected": 0,
            "students": [],
        }

    results = []

    # ============================================================
    # 9. PROCESS EVERY DETECTED FACE
    # ============================================================
    for live_embedding in face_embeddings:
        match = engine.recognize(live_embedding)

        # --------------------------------------------------------
        # UNKNOWN FACE
        # --------------------------------------------------------
        if match is None:
            results.append(
                {
                    "recognized": False,
                    "student": None,
                    "distance": None,
                    "attendance_marked": False,
                    "message": "Unknown face",
                }
            )
            continue

        student_id = match["student_id"]

        # --------------------------------------------------------
        # DUPLICATE IN CURRENT REQUEST
        # --------------------------------------------------------
        if engine.already_marked(student_id):
            results.append(
                {
                    "recognized": True,
                    "student": {
                        "id": student_id,
                        "name": match["name"],
                    },
                    "distance": match["distance"],
                    "attendance_marked": False,
                    "message": "Already processed",
                }
            )
            continue

        # --------------------------------------------------------
        # CHECK DATABASE FOR EXISTING ATTENDANCE
        # --------------------------------------------------------
        existing_record = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.session_id == session_id,
                AttendanceRecord.student_id == student_id,
            )
            .first()
        )

        # --------------------------------------------------------
        # RECORD ALREADY EXISTS
        # --------------------------------------------------------
        if existing_record:
            if existing_record.status == "PRESENT":
                engine.mark_as_processed(student_id)
                results.append(
                    {
                        "recognized": True,
                        "student": {
                            "id": student_id,
                            "name": match["name"],
                        },
                        "distance": match["distance"],
                        "attendance_marked": False,
                        "message": "Already marked present",
                    }
                )
                continue

            # Student exists but status is ABSENT
            existing_record.status = "PRESENT"
            existing_record.marked_at = datetime.now(timezone.utc)

            engine.mark_as_processed(student_id)
            results.append(
                {
                    "recognized": True,
                    "student": {
                        "id": student_id,
                        "name": match["name"],
                    },
                    "distance": match["distance"],
                    "attendance_marked": True,
                    "message": "Attendance marked successfully",
                }
            )
            continue

        # --------------------------------------------------------
        # NO ATTENDANCE RECORD EXISTS
        # --------------------------------------------------------
        attendance_record = AttendanceRecord(
            session_id=session_id,
            student_id=student_id,
            status="PRESENT",
            marked_at=datetime.now(timezone.utc),
        )

        db.add(attendance_record)
        engine.mark_as_processed(student_id)

        results.append(
            {
                "recognized": True,
                "student": {
                    "id": student_id,
                    "name": match["name"],
                },
                "distance": match["distance"],
                "attendance_marked": True,
                "message": "Attendance marked successfully",
            }
        )

    # ============================================================
    # 10. COMMIT ALL ATTENDANCE RECORDS
    # ============================================================
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to save attendance",
        )

    # ============================================================
    # 11. RETURN RESULT
    # ============================================================
    return {
        "session_id": session_id,
        "recognized": any(result["recognized"] for result in results),
        "faces_detected": len(face_embeddings),
        "students": results,
    }


@router.get(
    "/sessions/history"
)
def get_attendance_session_history(
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):
    # --------------------------------------------------------
    # 1. Get completed sessions for current teacher
    # --------------------------------------------------------

    sessions = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.teacher_id
            == current_teacher.id,
            AttendanceSession.status == "COMPLETED",
        )
        .order_by(
            AttendanceSession.started_at.desc()
        )
        .all()
    )

    # --------------------------------------------------------
    # 2. Build session history
    # --------------------------------------------------------

    session_history = []

    for session in sessions:

        # ----------------------------------------------------
        # Get total students
        # ----------------------------------------------------

        total_students = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.session_id
                == session.id
            )
            .count()
        )

        # ----------------------------------------------------
        # Get present students
        # ----------------------------------------------------

        present_students = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.session_id
                == session.id,
                AttendanceRecord.status == "PRESENT",
            )
            .count()
        )

        # ----------------------------------------------------
        # Get absent students
        # ----------------------------------------------------

        absent_students = (
            db.query(AttendanceRecord)
            .filter(
                AttendanceRecord.session_id
                == session.id,
                AttendanceRecord.status == "ABSENT",
            )
            .count()
        )

        # ----------------------------------------------------
        # Add session
        # ----------------------------------------------------

        session_history.append(
            {
                "id": session.id,
                "status": session.status,
                "started_at": session.started_at,
                "ended_at": session.ended_at,
                "statistics": {
                    "total_students": total_students,
                    "present": present_students,
                    "absent": absent_students,
                },
            }
        )

    # --------------------------------------------------------
    # 3. Return history
    # --------------------------------------------------------

    return {
        "teacher": {
            "id": current_teacher.id,
            "name": current_teacher.name,
        },
        "total_sessions": len(session_history),
        "sessions": session_history,
    }


@router.get(
    "/sessions/{session_id}"
)
def get_attendance_session_details(
    session_id: int,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find session belonging to current teacher
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. Session not found
    # --------------------------------------------------------

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # --------------------------------------------------------
    # 3. Get attendance records
    # --------------------------------------------------------

    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id
        )
        .order_by(
            AttendanceRecord.student_id
        )
        .all()
    )

    # --------------------------------------------------------
    # 4. Build student attendance list
    # --------------------------------------------------------

    students = []

    for record in records:

        student = record.student

        students.append(
            {
                "student_id": student.id,
                "name": student.name,
                "roll_number": student.roll_number,
                "email": student.email,
                "department": student.department,
                "semester": student.semester,
                "status": record.status,
                "marked_at": record.marked_at,
            }
        )

    # --------------------------------------------------------
    # 5. Calculate statistics
    # --------------------------------------------------------

    total_students = len(students)

    present_students = sum(
        1
        for student in students
        if student["status"] == "PRESENT"
    )

    absent_students = sum(
        1
        for student in students
        if student["status"] == "ABSENT"
    )

    # --------------------------------------------------------
    # 6. Return session details
    # --------------------------------------------------------

    return {
        "teacher": {
            "id": current_teacher.id,
            "name": current_teacher.name,
        },

        "session": {
            "id": session.id,
            "status": session.status,
            "started_at": session.started_at,
            "ended_at": session.ended_at,
        },

        "statistics": {
            "total_students": total_students,
            "present": present_students,
            "absent": absent_students,
        },

        "students": students,
    }


@router.get(
    "/sessions/{session_id}/report"
)
def get_attendance_report(
    session_id: int,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find session belonging to current teacher
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. Session not found
    # --------------------------------------------------------

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # --------------------------------------------------------
    # 3. Get attendance records
    # --------------------------------------------------------

    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id
        )
        .order_by(
            AttendanceRecord.student_id
        )
        .all()
    )

    # --------------------------------------------------------
    # 4. Build report rows
    # --------------------------------------------------------

    attendance = []

    for index, record in enumerate(
        records,
        start=1
    ):

        student = record.student

        attendance.append(
            {
                "s_no": index,
                "student_id": student.id,
                "name": student.name,
                "roll_number": student.roll_number,
                "email": student.email,
                "department": student.department,
                "semester": student.semester,
                "status": record.status,
                "marked_at": record.marked_at,
            }
        )

    # --------------------------------------------------------
    # 5. Calculate statistics
    # --------------------------------------------------------

    total_students = len(attendance)

    present_students = sum(
        1
        for record in attendance
        if record["status"] == "PRESENT"
    )

    absent_students = sum(
        1
        for record in attendance
        if record["status"] == "ABSENT"
    )

    attendance_percentage = (
        (
            present_students
            / total_students
        ) * 100
        if total_students > 0
        else 0
    )

    # --------------------------------------------------------
    # 6. Return report
    # --------------------------------------------------------

    return {
        "report": {
            "session_id": session.id,
            "session_status": session.status,
            "started_at": session.started_at,
            "ended_at": session.ended_at,
        },

        "teacher": {
            "id": current_teacher.id,
            "name": current_teacher.name,
        },

        "statistics": {
            "total_students": total_students,
            "present": present_students,
            "absent": absent_students,
            "attendance_percentage": round(
                attendance_percentage,
                2
            ),
        },

        "attendance": attendance,
    }

@router.get(
    "/sessions/{session_id}/report/csv"
)
def download_attendance_csv(
    session_id: int,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find session belonging to current teacher
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. Session not found
    # --------------------------------------------------------

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # --------------------------------------------------------
    # 3. Get attendance records
    # --------------------------------------------------------

    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id
        )
        .order_by(
            AttendanceRecord.student_id
        )
        .all()
    )

    # --------------------------------------------------------
    # 4. Create CSV in memory
    # --------------------------------------------------------

    output = io.StringIO()

    writer = csv.writer(output)

    # --------------------------------------------------------
    # 5. Report information
    # --------------------------------------------------------

    writer.writerow([
        "Attendance Report"
    ])

    writer.writerow([
        "Session ID",
        session.id
    ])

    writer.writerow([
        "Teacher",
        current_teacher.name
    ])

    writer.writerow([
        "Session Status",
        session.status
    ])

    writer.writerow([
        "Started At",
        session.started_at
    ])

    writer.writerow([
        "Ended At",
        session.ended_at
    ])

    writer.writerow([])

    # --------------------------------------------------------
    # 6. Table header
    # --------------------------------------------------------

    writer.writerow([
        "S.No",
        "Student ID",
        "Name",
        "Roll Number",
        "Email",
        "Department",
        "Semester",
        "Status",
        "Marked At",
    ])

    # --------------------------------------------------------
    # 7. Student attendance data
    # --------------------------------------------------------

    for index, record in enumerate(
        records,
        start=1
    ):

        student = record.student

        writer.writerow([
            index,
            student.id,
            student.name,
            student.roll_number,
            student.email,
            student.department,
            student.semester,
            record.status,
            record.marked_at,
        ])

    # --------------------------------------------------------
    # 8. Statistics
    # --------------------------------------------------------

    total_students = len(records)

    present_students = sum(
        1
        for record in records
        if record.status == "PRESENT"
    )

    absent_students = sum(
        1
        for record in records
        if record.status == "ABSENT"
    )

    attendance_percentage = (
        (
            present_students
            / total_students
        ) * 100
        if total_students > 0
        else 0
    )

    writer.writerow([])

    writer.writerow([
        "Statistics"
    ])

    writer.writerow([
        "Total Students",
        total_students
    ])

    writer.writerow([
        "Present",
        present_students
    ])

    writer.writerow([
        "Absent",
        absent_students
    ])

    writer.writerow([
        "Attendance Percentage",
        f"{attendance_percentage:.2f}%"
    ])

    # --------------------------------------------------------
    # 9. Move cursor to beginning
    # --------------------------------------------------------

    output.seek(0)

    # --------------------------------------------------------
    # 10. Create filename
    # --------------------------------------------------------

    filename = (
        f"attendance_session_{session.id}.csv"
    )

    # --------------------------------------------------------
    # 11. Return CSV file
    # --------------------------------------------------------

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )

@router.get(
    "/sessions/{session_id}/report/excel"
)
def download_attendance_excel(
    session_id: int,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(
        get_current_teacher
    ),
):

    # --------------------------------------------------------
    # 1. Find session belonging to current teacher
    # --------------------------------------------------------

    session = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.id == session_id,
            AttendanceSession.teacher_id
            == current_teacher.id,
        )
        .first()
    )

    # --------------------------------------------------------
    # 2. Session not found
    # --------------------------------------------------------

    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance session not found",
        )

    # --------------------------------------------------------
    # 3. Get attendance records
    # --------------------------------------------------------

    records = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.session_id
            == session.id
        )
        .order_by(
            AttendanceRecord.student_id
        )
        .all()
    )

    # --------------------------------------------------------
    # 4. Calculate statistics
    # --------------------------------------------------------

    total_students = len(records)

    present_students = sum(
        1
        for record in records
        if record.status == "PRESENT"
    )

    absent_students = sum(
        1
        for record in records
        if record.status == "ABSENT"
    )

    attendance_percentage = (
        (
            present_students
            / total_students
        ) * 100
        if total_students > 0
        else 0
    )

    # --------------------------------------------------------
    # 5. Create workbook
    # --------------------------------------------------------

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Attendance Report"

    # --------------------------------------------------------
    # 6. Report title
    # --------------------------------------------------------

    worksheet["A1"] = "Attendance Report"

    worksheet["A1"].font = Font(
        bold=True,
        size=16,
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    # Merge title
    worksheet.merge_cells(
        "A1:I1"
    )

    # --------------------------------------------------------
    # 7. Session information
    # --------------------------------------------------------

    worksheet["A3"] = "Session ID"
    worksheet["B3"] = session.id

    worksheet["A4"] = "Teacher"
    worksheet["B4"] = current_teacher.name

    worksheet["A5"] = "Session Status"
    worksheet["B5"] = session.status

    worksheet["A6"] = "Started At"
    worksheet["B6"] = session.started_at

    worksheet["A7"] = "Ended At"
    worksheet["B7"] = session.ended_at

    # Make labels bold
    for row in range(3, 8):

        worksheet[
            f"A{row}"
        ].font = Font(
            bold=True
        )

    # --------------------------------------------------------
    # 8. Table header
    # --------------------------------------------------------

    header_row = 9

    headers = [
        "S.No",
        "Student ID",
        "Name",
        "Roll Number",
        "Email",
        "Department",
        "Semester",
        "Status",
        "Marked At",
    ]

    for column, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column,
            value=header,
        )

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # --------------------------------------------------------
    # 9. Attendance records
    # --------------------------------------------------------

    for index, record in enumerate(
        records,
        start=1
    ):

        student = record.student

        row = header_row + index

        values = [
            index,
            student.id,
            student.name,
            student.roll_number,
            student.email,
            student.department,
            student.semester,
            record.status,
            record.marked_at,
        ]

        for column, value in enumerate(
            values,
            start=1
        ):

            worksheet.cell(
                row=row,
                column=column,
                value=value,
            )

    # --------------------------------------------------------
    # 10. Statistics
    # --------------------------------------------------------

    statistics_row = (
        header_row
        + total_students
        + 3
    )

    worksheet.cell(
        row=statistics_row,
        column=1,
        value="Statistics",
    ).font = Font(
        bold=True,
        size=13,
    )

    worksheet.cell(
        row=statistics_row + 1,
        column=1,
        value="Total Students",
    )

    worksheet.cell(
        row=statistics_row + 1,
        column=2,
        value=total_students,
    )

    worksheet.cell(
        row=statistics_row + 2,
        column=1,
        value="Present",
    )

    worksheet.cell(
        row=statistics_row + 2,
        column=2,
        value=present_students,
    )

    worksheet.cell(
        row=statistics_row + 3,
        column=1,
        value="Absent",
    )

    worksheet.cell(
        row=statistics_row + 3,
        column=2,
        value=absent_students,
    )

    worksheet.cell(
        row=statistics_row + 4,
        column=1,
        value="Attendance Percentage",
    )

    worksheet.cell(
        row=statistics_row + 4,
        column=2,
        value=(
            present_students / total_students
            if total_students > 0
            else 0
        ),
    )

    worksheet.cell(
        row=statistics_row + 4,
        column=2,
    ).number_format = "0.00%"

    # --------------------------------------------------------
    # 11. Bold statistics labels
    # --------------------------------------------------------

    for row in range(
        statistics_row + 1,
        statistics_row + 5,
    ):

        worksheet.cell(
            row=row,
            column=1,
        ).font = Font(
            bold=True
        )

    # --------------------------------------------------------
    # 12. Adjust column widths
    # --------------------------------------------------------

    column_widths = {
        "A": 10,
        "B": 14,
        "C": 25,
        "D": 18,
        "E": 32,
        "F": 35,
        "G": 12,
        "H": 15,
        "I": 25,
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # --------------------------------------------------------
    # 13. Freeze attendance header
    # --------------------------------------------------------

    worksheet.freeze_panes = "A10"

    # --------------------------------------------------------
    # 14. Save workbook to memory
    # --------------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    # --------------------------------------------------------
    # 15. Generate filename
    # --------------------------------------------------------

    filename = (
        f"attendance_session_{session.id}.xlsx"
    )

    # --------------------------------------------------------
    # 16. Return Excel file
    # --------------------------------------------------------

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )